import os

# import PyPDF2
from classic.components import component
from openpyxl import load_workbook

from pydantic import validate_arguments

from magnit.application import interfaces, errors

from magnit.application.services.join_point import join_point

NOT_TONAR_WEIGHING_ACT_TEMPLATE_PATH = 'weighing_act.xlsx'
TONAR_WEIGHING_ACT_TEMPLATE_PATH = 'weighing_act_tonar.xlsx'
TONAR_INVOICE_TEMPLATE_PATH = 'transport_invoice.xlsx'

FILE_TEMP = 'doc_temp.xlsx'  # TODO remove


@component
class Doc:
    """
    Класс Документы
    """
    visits_repo: interfaces.VisitRepo
    template_dir: str

    @join_point
    @validate_arguments
    def get_weighing_act(self, visit_id: int):

        # ../api/docs/weighing_act&visit_id=3

        visit = self.visits_repo.get_by_id(visit_id)
        if visit is None:
            raise errors.VisitIDNotExistError(visit_id=visit_id)

        template_path = NOT_TONAR_WEIGHING_ACT_TEMPLATE_PATH
        if visit.permit.is_tonar:
            template_path = TONAR_WEIGHING_ACT_TEMPLATE_PATH

        file_path = os.path.join(self.template_dir, template_path)
        wb = load_workbook(file_path)

        cell_map = {
            'polygon': ['A1', 'A17', 'A33'],
            'invoice_num': ['D3', 'D19', 'D35'],
            'checked_out': ['F3', 'F19', 'F35'],
            'transporter': ['D6', 'D22', 'D38'],
            'vehicle_model': ['D7', 'D23', 'D39'],
            'vehicle_number': ['D8', 'D24', 'D40'],
            'permit_number': ['A11', 'A27', 'A43'],
            'weight_in': ['D11', 'D27', 'D43'],
            'weight_out': ['E11', 'E27', 'E43'],
            'netto': ['G11', 'G27', 'G43'],
        }

        permit_number = visit.permit.id
        polygon = visit.polygon.name
        weight_in = visit.weight_in
        weight_out = visit.weight_out
        netto = visit.netto
        invoice_num = visit.invoice_num
        transporter = visit.permit.contragent.name
        vehicle_model = visit.permit.vehicle.model.model
        vehicle_number = visit.permit.vehicle.reg_number
        checked_out = visit.checked_out

        ws = wb.active

        self._fill_cells(ws, polygon, cell_map['polygon'])
        self._fill_cells(ws, invoice_num, cell_map['invoice_num'])
        self._fill_cells(ws, checked_out, cell_map['checked_out'])
        self._fill_cells(ws, transporter, cell_map['transporter'])
        self._fill_cells(ws, vehicle_model, cell_map['vehicle_model'])
        self._fill_cells(ws, vehicle_number, cell_map['vehicle_number'])
        self._fill_cells(ws, permit_number, cell_map['permit_number'])
        self._fill_cells(ws, weight_in, cell_map['weight_in'])
        self._fill_cells(ws, weight_out, cell_map['weight_out'])
        self._fill_cells(ws, netto, cell_map['netto'])

        wb.save(FILE_TEMP)  # TODO save as PDF

        file_path = os.path.join(FILE_TEMP)     # TODO temporary file dir
        return file_path

    def _fill_cells(self, ws, value, cells):
        for c in cells:
            ws[c] = value

    @join_point
    @validate_arguments
    def get_transport_invoice(self, visit_id: int):

        visit = self.visits_repo.get_by_id(visit_id)
        if visit is None:
            raise errors.VisitIDNotExistError(visit_id=visit_id)

        template_path = os.path.join(
            self.template_dir,
            TONAR_INVOICE_TEMPLATE_PATH,
        )

        wb = load_workbook(template_path)
        cell_map = {
            'polygon': 'B60',
            'polygon_owner': 'BP15',
            'invoice_num': 'AC9',
            'invoice_date': 'G9',
            'transporter': 'B44',
            'transporter_details': 'B101',
            'vehicle_model': 'B47',
            'vehicle_number': 'BF47',
            'driver': 'BF44',
            'destination': 'B83',
            'cargo_description': 'B27',
        }

        polygon = visit.polygon.full_name
        polygon_owner = visit.polygon.owner.name
        invoice_num = visit.invoice_num
        invoice_date = visit.invoice_date
        transporter = '%s ИНН: %s, КПП: %s' % (
            visit.permit.contragent.name,
            visit.permit.contragent.inn,
            visit.permit.contragent.kpp,
        )
        transporter_details = '%s ,%s' % (
            transporter,
            visit.permit.contragent.address,
        )
        vehicle_model = visit.permit.vehicle.model.model
        vehicle_number = visit.permit.vehicle.reg_number
        driver = '%s %s' % (
            visit.driver.full_name,
            visit.driver.phone_print,
        )
        cargo_description = 'Объем груза %s м3, вес согласно Акту' \
                            ' взвешивания № %s от %s' % (
            visit.permit.vehicle.body_volume,
            invoice_num,
            invoice_date,
        )
        destination = visit.destination.full_name

        ws = wb.active
        ws[cell_map.get('polygon')] = polygon
        ws[cell_map.get('polygon_owner')] = polygon_owner
        ws[cell_map.get('invoice_num')] = invoice_num
        ws[cell_map.get('invoice_date')] = invoice_date
        ws[cell_map.get('transporter')] = transporter
        ws[cell_map.get('transporter_details')] = transporter_details
        ws[cell_map.get('vehicle_model')] = vehicle_model
        ws[cell_map.get('vehicle_number')] = vehicle_number
        ws[cell_map.get('driver')] = driver
        ws[cell_map.get('destination')] = destination
        ws[cell_map.get('cargo_description')] = cargo_description

        wb.save(FILE_TEMP)

    # @join_point
    # @validate_arguments
    # def get_xlsx_to_pdf(self):
    #     template_path = TONAR_WEIGHING_ACT_TEMPLATE_PATH
    #     file_path = os.path.join(self.template_dir, template_path)
    #     wb = Workbook(file_path)
    #     wb.Save("Output.pdf")
    #     jpype.shutdownGuiEnvironment()
    #     ws = wb.active
    #     pw = PdfWriter('from_xlsx.pdf')
    #     pw.setFont('Courier', 12)
    #     pw.setHeader('XLSXtoPDF.py - convert XLSX data to PDF')
    #     pw.setFooter('Generated using openpyxl and xtopdf')
    #
    #     ws_range = ws.iter_rows('A1:H13')
    #     for row in ws_range:
    #         s = ''
    #         for cell in row:
    #             if cell.value is None:
    #                 s += ' ' * 11
    #             else:
    #                 s += str(cell.value).rjust(10) + ' '
    #         pw.writeLine(s)
    #     pw.savePage()
    #     pw.close()
